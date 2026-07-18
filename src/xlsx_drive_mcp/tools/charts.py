import io
from typing import Optional

import openpyxl
from fastmcp import FastMCP
from googleapiclient.discovery import Resource
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import range_boundaries, get_column_letter

from ..drive import download_xlsx, upload_xlsx, validate_cell_range

CHART_TYPES = {"bar", "line", "pie", "scatter"}

_CHART_SUBTYPES = {
    "bar": {"col", "bar"},
    "line": {"line", "smooth"},
    "pie": {"pie"},
    "scatter": {"marker", "smooth_line", "straight_line"},
}


def _build_chart(ws, chart_type: str, chart_subtype: str, values_range: str,
                 category_range: Optional[str], series_names: Optional[list],
                 title: Optional[str]):
    if chart_type not in CHART_TYPES:
        raise ValueError(f"Invalid chart_type '{chart_type}'. Must be one of: {sorted(CHART_TYPES)}")

    min_col, min_row, max_col, max_row = range_boundaries(values_range.upper())

    if chart_type == "bar":
        chart = BarChart()
        chart.type = chart_subtype if chart_subtype in {"col", "bar"} else "col"
    elif chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    elif chart_type == "scatter":
        chart = ScatterChart()

    if title:
        chart.title = title

    # Add data series -- each column in values_range is one series
    for col_idx in range(min_col, max_col + 1):
        values_ref = Reference(ws, min_col=col_idx, min_row=min_row, max_row=max_row)
        series = Series(values_ref)
        if series_names and isinstance(series_names, list):
            s_idx = col_idx - min_col
            if s_idx < len(series_names):
                series.title = SeriesLabel(v=series_names[s_idx])
        elif series_names and isinstance(series_names, str):
            # series_names is a cell range like "B1:D1"
            sc, sr, ec, er = range_boundaries(series_names.upper())
            name_ref = Reference(ws, min_col=sc + (col_idx - min_col), min_row=sr)
            series.title = name_ref
        chart.series.append(series)

    if category_range:
        cat_min_col, cat_min_row, _, cat_max_row = range_boundaries(category_range.upper())
        categories = Reference(ws, min_col=cat_min_col, min_row=cat_min_row, max_row=cat_max_row)
        chart.set_categories(categories)

    return chart


def _place_chart(ws, chart, anchor_cell: str) -> None:
    ws.add_chart(chart, anchor_cell.upper())


def _default_anchor(values_range: str) -> str:
    """Place chart one column to the right of the data range."""
    _, _, max_col, min_row = range_boundaries(values_range.upper())
    return f"{get_column_letter(max_col + 2)}{min_row}"


def register_tools(mcp: FastMCP, service: Resource) -> None:

    @mcp.tool()
    def create_chart(
        file_id: str,
        sheet_name: str,
        chart_type: str,
        values_range: str,
        chart_subtype: Optional[str] = None,
        category_range: Optional[str] = None,
        series_names: Optional[list] = None,
        title: Optional[str] = None,
        anchor_cell: Optional[str] = None,
    ) -> dict:
        """Create a chart in an xlsx file on Google Drive.

        chart_type: 'bar', 'line', 'pie', 'scatter'
        chart_subtype: bar='col'(default)/'bar'; line='line'(default)/'smooth';
                       pie='pie'; scatter='marker'(default)/'smooth_line'/'straight_line'
        values_range: cell range of data values, e.g. 'B2:D10'. Columns = series.
        category_range: cell range of X-axis/category labels, e.g. 'A2:A10'.
        series_names: list of series label strings, or a cell range like 'B1:D1'.
        anchor_cell: top-left cell for chart placement, e.g. 'E2'.
        """
        validate_cell_range(values_range)
        if category_range:
            validate_cell_range(category_range)

        subtypes = _CHART_SUBTYPES.get(chart_type, set())
        if chart_subtype is None:
            chart_subtype = next(iter(subtypes))  # default = first
        elif chart_subtype not in subtypes:
            raise ValueError(
                f"Invalid chart_subtype '{chart_subtype}' for chart_type '{chart_type}'. "
                f"Valid options: {sorted(subtypes)}"
            )

        if anchor_cell is None:
            anchor_cell = _default_anchor(values_range)

        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

        ws = wb[sheet_name]
        chart = _build_chart(ws, chart_type, chart_subtype, values_range,
                              category_range, series_names, title)
        _place_chart(ws, chart, anchor_cell)

        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"sheet_name": sheet_name, "anchor_cell": anchor_cell, "chart_type": chart_type}
