import io
from typing import Optional

import openpyxl
from fastmcp import FastMCP
from googleapiclient.discovery import Resource
from openpyxl.utils import range_boundaries, get_column_letter

from ..drive import download_xlsx, upload_xlsx, validate_cell_range


def _read_range(wb: openpyxl.Workbook, sheet_name: str, cell_range: str) -> dict:
    ws = wb[sheet_name]  # raises KeyError if not found
    if ":" in cell_range:
        rows = list(ws[cell_range.upper()])
        values = [[cell.value for cell in row] for row in rows]
    else:
        cell = ws[cell_range.upper()]
        values = [[cell.value]]
    return {"values": values, "cell_range": cell_range}


def _write_range(wb: openpyxl.Workbook, sheet_name: str, cell_range: str, values: list) -> None:
    ws = wb[sheet_name]
    # Determine anchor from top-left of cell_range
    anchor = cell_range.upper().split(":")[0]
    min_col, min_row, _, _ = range_boundaries(anchor + ":" + anchor)
    for r_offset, row_values in enumerate(values):
        for c_offset, value in enumerate(row_values):
            ws.cell(row=min_row + r_offset, column=min_col + c_offset, value=value)


def _find_last_data_row(ws) -> int:
    """Scan from bottom to find the last row that has at least one non-None value."""
    last_row = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                last_row = max(last_row, cell.row)
    return last_row


def _append_rows(wb: openpyxl.Workbook, sheet_name: str, rows: list) -> None:
    ws = wb[sheet_name]
    last_row = _find_last_data_row(ws)
    for r_offset, row_values in enumerate(rows):
        for c_offset, value in enumerate(row_values):
            ws.cell(row=last_row + 1 + r_offset, column=1 + c_offset, value=value)


def _clear_range(wb: openpyxl.Workbook, sheet_name: str, cell_range: str) -> None:
    ws = wb[sheet_name]
    merged = {str(r) for r in ws.merged_cells.ranges}  # noqa: F841 — kept per spec
    min_col, min_row, max_col, max_row = range_boundaries(cell_range.upper())
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            addr = f"{get_column_letter(c)}{r}"
            for merge_range in ws.merged_cells.ranges:
                if ws[addr].coordinate in [
                    ws.cell(row=mr, column=mc).coordinate
                    for mr in range(merge_range.min_row, merge_range.max_row + 1)
                    for mc in range(merge_range.min_col, merge_range.max_col + 1)
                ]:
                    raise ValueError(
                        f"Cell {addr} is part of a merged region {merge_range}. "
                        f"Call unmerge_cells first."
                    )
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None


def register_tools(mcp: FastMCP, service: Resource) -> None:

    @mcp.tool()
    def read_range(
        file_id: str,
        sheet_name: str,
        cell_range: str,
        data_only: bool = True,
    ) -> dict:
        """Read cell values from a range in an xlsx file on Google Drive.

        When data_only=True (default), formula cells return their last cached computed
        value from when Excel/Sheets last saved the file. When data_only=False, formula
        strings are returned instead. These modes are mutually exclusive in openpyxl.

        Args:
            cell_range: A1-notation range such as 'A1' or 'A1:D10'.
        """
        validate_cell_range(cell_range)
        content, _ = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=data_only)
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return _read_range(wb, sheet_name, cell_range)

    @mcp.tool()
    def write_range(
        file_id: str,
        sheet_name: str,
        cell_range: str,
        values: list,
    ) -> dict:
        """Write a 2D array of values to an xlsx file on Google Drive.

        cell_range sets the top-left anchor. The values array determines the write
        extent. Cells beyond the values array are not cleared. Formula strings
        (starting with '=') are written as formulas.

        WARNING: Writing over cells that contain formulas replaces them with
        the static values provided.
        """
        validate_cell_range(cell_range)
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        _write_range(wb, sheet_name, cell_range, values)
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"written": True, "cell_range": cell_range, "rows": len(values)}

    @mcp.tool()
    def append_rows(file_id: str, sheet_name: str, rows: list) -> dict:
        """Append rows below the last non-empty row in a sheet.

        Uses an actual data scan to find the last row -- not openpyxl's max_row
        property, which is unreliable after cells have been cleared.
        """
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        _append_rows(wb, sheet_name, rows)
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"appended_rows": len(rows)}

    @mcp.tool()
    def clear_range(file_id: str, sheet_name: str, cell_range: str) -> dict:
        """Clear cell values in a range, preserving formatting.

        Raises an error if the range intersects a merged region (call
        unmerge_cells first).
        """
        validate_cell_range(cell_range)
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        _clear_range(wb, sheet_name, cell_range)
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"cleared": True, "cell_range": cell_range}
