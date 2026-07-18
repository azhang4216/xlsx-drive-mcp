import io
import re
from typing import Optional

import openpyxl
from fastmcp import FastMCP
from googleapiclient.discovery import Resource

from ..drive import (
    XLSX_MIME_TYPE,
    download_xlsx,
)

_SINGLE_QUOTE_RE = re.compile(r"'")


def _build_list_query(folder_id: Optional[str], name_contains: Optional[str]) -> str:
    parts = [f"mimeType='{XLSX_MIME_TYPE}'", "trashed=false"]
    if folder_id:
        parts.append(f"'{folder_id}' in parents")
    if name_contains:
        escaped = _SINGLE_QUOTE_RE.sub("\\'", name_contains)
        parts.append(f"name contains '{escaped}'")
    return " and ".join(parts)


def _scan_sheet_dimensions(ws) -> tuple[int, int]:
    """Return (last_data_row, last_data_col) by scanning actual cell values."""
    max_row = 0
    max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    return max_row, max_col


def _get_xlsx_info_from_wb(wb: openpyxl.Workbook) -> dict:
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows, cols = _scan_sheet_dimensions(ws)
        sheets.append({"name": name, "rows": rows, "cols": cols})
    named_ranges = list(wb.defined_names.keys())
    return {"sheets": sheets, "named_ranges": named_ranges}


def register_tools(mcp: FastMCP, service: Resource) -> None:

    @mcp.tool()
    def list_xlsx_files(
        folder_id: Optional[str] = None,
        name_contains: Optional[str] = None,
        page_token: Optional[str] = None,
    ) -> dict:
        """List xlsx files on Google Drive.

        Args:
            folder_id: Restrict search to this Drive folder ID.
            name_contains: Filter by filename substring (safe, not raw query).
            page_token: Token from a previous response to fetch the next page.

        Returns dict with 'files' list and optional 'next_page_token'.
        """
        q = _build_list_query(folder_id, name_contains)
        params = dict(
            q=q,
            fields="nextPageToken,files(id,name,modifiedTime,size)",
            pageSize=100,
        )
        if page_token:
            params["pageToken"] = page_token
        result = service.files().list(**params).execute()
        files = [
            {
                "file_id": f["id"],
                "name": f["name"],
                "modified_time": f.get("modifiedTime"),
                "size_bytes": int(f.get("size", 0)),
            }
            for f in result.get("files", [])
        ]
        out = {"files": files}
        if "nextPageToken" in result:
            out["next_page_token"] = result["nextPageToken"]
        return out

    @mcp.tool()
    def get_xlsx_info(file_id: str) -> dict:
        """Get sheet names, data dimensions, and named ranges for an xlsx file.

        Dimensions are determined by scanning actual data, not openpyxl's
        max_row/max_column (which are unreliable after cells are cleared).
        """
        content, _ = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        return _get_xlsx_info_from_wb(wb)

    @mcp.tool()
    def create_xlsx(name: str, folder_id: Optional[str] = None) -> dict:
        """Create a new blank xlsx file on Google Drive. Returns the file_id."""
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        metadata = {"name": name, "mimeType": XLSX_MIME_TYPE}
        if folder_id:
            metadata["parents"] = [folder_id]
        from googleapiclient.http import MediaIoBaseUpload
        media = MediaIoBaseUpload(io.BytesIO(buf.getvalue()), mimetype=XLSX_MIME_TYPE)
        result = service.files().create(
            body=metadata,
            media_body=media,
            fields="id,name",
        ).execute()
        return {"file_id": result["id"], "name": result["name"]}

    @mcp.tool()
    def delete_xlsx(file_id: str) -> dict:
        """Move an xlsx file to Drive trash.

        WARNING: This is non-reversible from the agent's perspective.
        The user has 30 days to recover the file from the Drive trash UI.
        Use with caution.
        """
        service.files().update(fileId=file_id, body={"trashed": True}).execute()
        return {"deleted": True, "file_id": file_id}
