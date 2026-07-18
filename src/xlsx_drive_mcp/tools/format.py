import io
import re
from copy import copy
from typing import Optional, Union

import openpyxl
from fastmcp import FastMCP
from googleapiclient.discovery import Resource
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries

from ..drive import download_xlsx, upload_xlsx, validate_cell_range

_HEX6_RE = re.compile(r"^[0-9A-Fa-f]{6}$")
_HEX8_RE = re.compile(r"^[0-9A-Fa-f]{8}$")


def _parse_color_input(color: str) -> str:
    """Normalize color to 8-digit ARGB hex string."""
    if _HEX8_RE.match(color):
        return color.upper()
    if _HEX6_RE.match(color):
        return "FF" + color.upper()
    raise ValueError(f"Invalid color '{color}'. Use RRGGBB or FFRRGGBB hex format.")


def _parse_border_input(border) -> dict:
    """Normalize border param to {top, bottom, left, right} dict."""
    if isinstance(border, str):
        return {"top": border, "bottom": border, "left": border, "right": border}
    return border  # already a dict


def _expand_columns(columns) -> list:
    """Expand 'A', ['A','B'], or 'A:C' to list of column letters."""
    if isinstance(columns, str) and ":" in columns:
        start, end = columns.split(":")
        s = column_index_from_string(start.upper())
        e = column_index_from_string(end.upper())
        return [get_column_letter(i) for i in range(s, e + 1)]
    if isinstance(columns, str):
        return [columns.upper()]
    return [c.upper() for c in columns]


def _expand_rows(rows) -> list:
    """Expand 3, [1,2,3], or '2:4' to list of row ints."""
    if isinstance(rows, int):
        return [rows]
    if isinstance(rows, str) and ":" in rows:
        start, end = rows.split(":")
        return list(range(int(start), int(end) + 1))
    return list(rows)


def _get_color_str(color_obj) -> Optional[str]:
    """Extract ARGB hex string from an openpyxl Color object, or None."""
    if color_obj is None:
        return None
    if hasattr(color_obj, "type"):
        if color_obj.type == "rgb" and color_obj.rgb not in ("00000000", "FFFFFFFF", "FF000000") or color_obj.rgb:
            return color_obj.rgb if color_obj.rgb != "00000000" else None
    return None


def _read_cell_format(cell, ws) -> dict:
    """Extract formatting dict from a single openpyxl cell."""
    font = cell.font or Font()
    fill = cell.fill
    border = cell.border or Border()

    bg_color = None
    if fill and fill.fill_type == "solid" and fill.fgColor:
        raw = fill.fgColor.rgb if hasattr(fill.fgColor, "rgb") else None
        if raw and raw not in ("00000000",):
            bg_color = raw

    font_color = None
    if font.color and hasattr(font.color, "rgb"):
        raw = font.color.rgb
        if raw and raw not in ("FF000000", "00000000"):
            font_color = raw
        elif raw == "FF000000":
            font_color = "FF000000"

    def side_style(side):
        return side.border_style if side and side.border_style else None

    merge_range = None
    merged = False
    for merge in ws.merged_cells.ranges:
        if (merge.min_row <= cell.row <= merge.max_row and
                merge.min_col <= cell.column <= merge.max_col):
            merged = True
            if cell.row == merge.min_row and cell.column == merge.min_col:
                merge_range = str(merge)
            break

    return {
        "address": cell.coordinate,
        "bold": font.bold or False,
        "italic": font.italic or False,
        "font_size": font.size,
        "font_color": font_color,
        "bg_color": bg_color,
        "number_format": cell.number_format,
        "border": {
            "top": side_style(border.top),
            "bottom": side_style(border.bottom),
            "left": side_style(border.left),
            "right": side_style(border.right),
        },
        "merged": merged,
        "merge_range": merge_range,
    }


def _apply_format_to_range(ws, cell_range: str, **kwargs) -> None:
    """Apply formatting kwargs to all cells in cell_range."""
    min_col, min_row, max_col, max_row = range_boundaries(cell_range.upper())
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if "bold" in kwargs or "italic" in kwargs or "font_size" in kwargs or "font_color" in kwargs:
                existing = copy(cell.font) if cell.font else Font()
                cell.font = Font(
                    bold=kwargs.get("bold", existing.bold),
                    italic=kwargs.get("italic", existing.italic),
                    size=kwargs.get("font_size", existing.size),
                    color=_parse_color_input(kwargs["font_color"]) if kwargs.get("font_color") else existing.color,
                )
            if "bg_color" in kwargs:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=_parse_color_input(kwargs["bg_color"]),
                )
            if "number_format" in kwargs:
                cell.number_format = kwargs["number_format"]
            if "border" in kwargs:
                parsed = _parse_border_input(kwargs["border"])

                def make_side(style_val):
                    if style_val == "none":
                        return Side(style=None)
                    if style_val is None:
                        return copy(getattr(cell.border, "top", Side()))  # unchanged handled per-side
                    return Side(style=style_val)

                existing_border = copy(cell.border) if cell.border else Border()
                cell.border = Border(
                    top=make_side(parsed.get("top")) if "top" in parsed else existing_border.top,
                    bottom=make_side(parsed.get("bottom")) if "bottom" in parsed else existing_border.bottom,
                    left=make_side(parsed.get("left")) if "left" in parsed else existing_border.left,
                    right=make_side(parsed.get("right")) if "right" in parsed else existing_border.right,
                )


def register_tools(mcp: FastMCP, service: Resource) -> None:

    @mcp.tool()
    def read_format(file_id: str, sheet_name: str, cell_range: str) -> dict:
        """Read formatting (bold, colors, borders, merge status) for a cell range.

        Colors are returned as 8-digit ARGB hex strings (e.g. 'FFFF0000' = opaque red).
        Transparent/unset colors are returned as null.
        """
        validate_cell_range(cell_range)
        content, _ = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        cells_fmt = []
        if ":" in cell_range:
            for row in ws[cell_range.upper()]:
                for cell in row:
                    cells_fmt.append(_read_cell_format(cell, ws))
        else:
            cells_fmt.append(_read_cell_format(ws[cell_range.upper()], ws))
        return {"cells": cells_fmt}

    @mcp.tool()
    def format_range(
        file_id: str,
        sheet_name: str,
        cell_range: str,
        bold: Optional[bool] = None,
        italic: Optional[bool] = None,
        font_size: Optional[float] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
        number_format: Optional[str] = None,
        border: Optional[Union[str, dict]] = None,
    ) -> dict:
        """Apply formatting to a cell range. Omitted params are left unchanged.

        border: 'thin'|'medium'|'thick'|'none' (all sides), or
                {"top": "thin", "bottom": "none", ...} for per-side control.
                null means leave unchanged; 'none' removes the border.
        Colors: 'RRGGBB' or 'FFRRGGBB' hex format.
        """
        validate_cell_range(cell_range)
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        kwargs = {k: v for k, v in {
            "bold": bold, "italic": italic, "font_size": font_size,
            "font_color": font_color, "bg_color": bg_color,
            "number_format": number_format, "border": border,
        }.items() if v is not None}
        _apply_format_to_range(wb[sheet_name], cell_range, **kwargs)
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"formatted": True, "cell_range": cell_range}

    @mcp.tool()
    def set_column_width(
        file_id: str, sheet_name: str, columns: Union[str, list], width: float
    ) -> dict:
        """Set column width. columns: 'A', ['A','B'], or 'A:C'."""
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        for col in _expand_columns(columns):
            ws.column_dimensions[col].width = width
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"set_columns": _expand_columns(columns), "width": width}

    @mcp.tool()
    def set_row_height(
        file_id: str, sheet_name: str, rows: Union[int, str, list], height: float
    ) -> dict:
        """Set row height. rows: 3, [1,2,3], or '2:4'."""
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
        for row in _expand_rows(rows):
            ws.row_dimensions[row].height = height
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"set_rows": _expand_rows(rows), "height": height}

    @mcp.tool()
    def merge_cells(file_id: str, sheet_name: str, cell_range: str) -> dict:
        """Merge a cell range."""
        validate_cell_range(cell_range)
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        wb[sheet_name].merge_cells(cell_range.upper())
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"merged": cell_range}

    @mcp.tool()
    def unmerge_cells(file_id: str, sheet_name: str, cell_range: str) -> dict:
        """Unmerge a previously merged cell range."""
        validate_cell_range(cell_range)
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        wb[sheet_name].unmerge_cells(cell_range.upper())
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"unmerged": cell_range}
