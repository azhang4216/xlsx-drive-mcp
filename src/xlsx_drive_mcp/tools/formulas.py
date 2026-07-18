import io
import re

import openpyxl
from fastmcp import FastMCP
from googleapiclient.discovery import Resource

from ..drive import download_xlsx, upload_xlsx
from .data import _write_range

BLOCKED_FUNCTIONS = {
    "WEBSERVICE", "FILTERXML", "IMPORTDATA", "IMPORTFEED",
    "IMPORTHTML", "IMPORTRANGE", "IMPORTXML",
}

_BLOCKED_RE = re.compile(
    r"\b(" + "|".join(BLOCKED_FUNCTIONS) + r")\s*\(",
    re.IGNORECASE,
)


def _validate_formula(formula: str) -> None:
    """Raise ValueError if formula is invalid or uses a blocked function."""
    if not formula.startswith("="):
        raise ValueError(
            f"Formula must start with '='. Got: '{formula[:30]}'"
        )
    match = _BLOCKED_RE.search(formula)
    if match:
        func_name = match.group(1).upper()
        raise ValueError(
            f"Formula uses blocked function '{func_name}', which makes external network "
            f"connections when the file is opened in Excel or Google Sheets. "
            f"Blocked functions: {sorted(BLOCKED_FUNCTIONS)}"
        )


def register_tools(mcp: FastMCP, service: Resource) -> None:

    @mcp.tool()
    def write_formula(file_id: str, sheet_name: str, cell: str, formula: str) -> dict:
        """Write a formula to a single cell in an xlsx file on Google Drive.

        Equivalent to write_range with a 1x1 values array. The formula must
        start with '='. Functions that make external network connections are
        blocked: WEBSERVICE, FILTERXML, IMPORTDATA, IMPORTFEED, IMPORTHTML,
        IMPORTRANGE, IMPORTXML.
        """
        _validate_formula(formula)
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        _write_range(wb, sheet_name, cell, [[formula]])
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"written": True, "cell": cell.upper(), "formula": formula}
