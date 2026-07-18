import io
from typing import Optional

import openpyxl
from fastmcp import FastMCP
from googleapiclient.discovery import Resource

from ..drive import download_xlsx, upload_xlsx


def _list_sheets(wb: openpyxl.Workbook) -> list[dict]:
    return [{"name": name, "index": i} for i, name in enumerate(wb.sheetnames)]


def _create_sheet(wb: openpyxl.Workbook, name: str, position: Optional[int]) -> dict:
    ws = wb.create_sheet(title=name, index=position)
    return {"name": ws.title, "index": wb.sheetnames.index(ws.title)}


def _delete_sheet(wb: openpyxl.Workbook, sheet_name: str) -> None:
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    if len(wb.sheetnames) == 1:
        raise ValueError("Cannot delete the only sheet in the workbook.")
    del wb[sheet_name]


def _rename_sheet(wb: openpyxl.Workbook, sheet_name: str, new_name: str) -> None:
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    if new_name in wb.sheetnames:
        raise ValueError(f"Sheet '{new_name}' already exists.")
    wb[sheet_name].title = new_name


def _copy_sheet(wb: openpyxl.Workbook, sheet_name: str, new_name: str) -> dict:
    """Duplicate sheet_name as new_name within the same workbook.

    WARNING: openpyxl's copy_worksheet silently drops charts, images,
    and some merged cell configurations from the copied sheet.
    """
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    if new_name in wb.sheetnames:
        raise ValueError(f"Sheet '{new_name}' already exists.")
    new_ws = wb.copy_worksheet(wb[sheet_name])
    new_ws.title = new_name
    return {"name": new_ws.title, "index": wb.sheetnames.index(new_ws.title)}


def register_tools(mcp: FastMCP, service: Resource) -> None:

    @mcp.tool()
    def list_sheets(file_id: str) -> list[dict]:
        """List all sheets in an xlsx file. Returns [{name, index}]."""
        if not file_id or not file_id.strip():
            raise ValueError("file_id must be a non-empty string")
        content, _ = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        return _list_sheets(wb)

    @mcp.tool()
    def create_sheet(file_id: str, name: str, position: Optional[int] = None) -> dict:
        """Add a new sheet to an xlsx file. Position is 0-based; defaults to last."""
        if not file_id or not file_id.strip():
            raise ValueError("file_id must be a non-empty string")
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        result = _create_sheet(wb, name, position)
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return result

    @mcp.tool()
    def delete_sheet(file_id: str, sheet_name: str) -> dict:
        """Remove a sheet from an xlsx file. Raises if it is the only sheet."""
        if not file_id or not file_id.strip():
            raise ValueError("file_id must be a non-empty string")
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        _delete_sheet(wb, sheet_name)
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"deleted": sheet_name}

    @mcp.tool()
    def rename_sheet(file_id: str, sheet_name: str, new_name: str) -> dict:
        """Rename a sheet in an xlsx file."""
        if not file_id or not file_id.strip():
            raise ValueError("file_id must be a non-empty string")
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        _rename_sheet(wb, sheet_name, new_name)
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return {"old_name": sheet_name, "new_name": wb[new_name].title}

    @mcp.tool()
    def copy_sheet(file_id: str, sheet_name: str, new_name: str) -> dict:
        """Duplicate a sheet within the same workbook.

        WARNING: openpyxl's copy_worksheet silently drops charts, images,
        and some merged cell configurations from the copied sheet.
        """
        if not file_id or not file_id.strip():
            raise ValueError("file_id must be a non-empty string")
        content, rev_id = download_xlsx(service, file_id)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        result = _copy_sheet(wb, sheet_name, new_name)
        buf = io.BytesIO()
        wb.save(buf)
        upload_xlsx(service, file_id, buf.getvalue(), rev_id)
        return result
