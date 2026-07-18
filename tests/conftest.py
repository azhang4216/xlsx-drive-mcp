import io
import pytest
import openpyxl
from unittest.mock import MagicMock


@pytest.fixture
def simple_wb_bytes() -> bytes:
    """Minimal xlsx: Sheet1 with A1=hello, B1=world, A2=42, B2=3.14."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["B1"] = "world"
    ws["A2"] = 42
    ws["B2"] = 3.14
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def mock_service():
    """Minimal Drive service mock. Tests configure return values as needed."""
    return MagicMock()


@pytest.fixture
def drive_store(simple_wb_bytes):
    """
    In-memory dict simulating Drive: {file_id: (content_bytes, revision_id)}.
    Returns the store dict; tests can add more files.
    """
    return {"file1": (simple_wb_bytes, "rev1")}


def make_download_fn(store: dict):
    from xlsx_drive_mcp.drive import FileTooLargeError, MAX_FILE_SIZE_BYTES

    def download(service, file_id: str) -> tuple[bytes, str]:
        if file_id not in store:
            raise KeyError(f"File {file_id!r} not in mock store")
        content, rev = store[file_id]
        if len(content) > MAX_FILE_SIZE_BYTES:
            raise FileTooLargeError("File too large")
        return content, rev

    return download


def make_upload_fn(store: dict):
    from xlsx_drive_mcp.drive import ConflictError

    def upload(service, file_id: str, content: bytes, revision_id: str) -> None:
        if file_id not in store:
            raise KeyError(f"File {file_id!r} not in mock store")
        _, current_rev = store[file_id]
        if current_rev != revision_id:
            raise ConflictError("Conflict")
        new_rev = f"rev{len(store[file_id][0]) + len(content)}"
        store[file_id] = (content, new_rev)

    return upload
