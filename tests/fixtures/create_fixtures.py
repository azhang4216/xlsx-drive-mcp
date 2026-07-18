"""Run once to generate fixture xlsx files: python tests/fixtures/create_fixtures.py"""
import io
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

FIXTURES = Path(__file__).parent


def make_merged_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Merged Header"
    ws.merge_cells("A1:C1")
    ws["A2"] = "Under merge"
    wb.save(FIXTURES / "merged_cells.xlsx")


def make_named_ranges():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i in range(1, 6):
        ws[f"A{i}"] = i * 10
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names["MyRange"] = DefinedName("MyRange", attr_text="Sheet1!$A$1:$A$5")
    wb.save(FIXTURES / "named_ranges.xlsx")


def make_formulas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(FIXTURES / "formulas.xlsx")


def make_multi_sheet():
    wb = openpyxl.Workbook()
    wb.active.title = "Alpha"
    wb.active["A1"] = "alpha data"
    wb.create_sheet("Beta")
    wb["Beta"]["A1"] = "beta data"
    wb.create_sheet("Gamma")
    wb["Gamma"]["A1"] = "gamma data"
    wb.save(FIXTURES / "multi_sheet.xlsx")


def make_formatting():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Bold Red"
    ws["A1"].font = Font(bold=True, color="FFFF0000")
    ws["B1"] = "Yellow BG"
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    ws["C1"] = "Bordered"
    thin = Side(style="thin")
    ws["C1"].border = Border(top=thin, bottom=thin, left=thin, right=thin)
    wb.save(FIXTURES / "formatting.xlsx")


if __name__ == "__main__":
    make_merged_cells()
    make_named_ranges()
    make_formulas()
    make_multi_sheet()
    make_formatting()
    print("Fixtures created.")
