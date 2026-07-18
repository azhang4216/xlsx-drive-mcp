# tests/integration/test_drive_roundtrip.py
"""
Integration tests against a real Google Drive file.
Requires:
  - DRIVE_TEST_FILE_ID env var: a real xlsx file_id on Drive
  - Credentials at ~/.google_workspace_mcp/credentials/{email}.json
    OR GOOGLE_CLIENT_ID + GOOGLE_CLIENT_SECRET with stored credentials

Run: python -m pytest tests/integration/ -v
"""
import io
import os
import openpyxl
import pytest

DRIVE_TEST_FILE_ID = os.environ.get("DRIVE_TEST_FILE_ID")

pytestmark = pytest.mark.skipif(
    not DRIVE_TEST_FILE_ID,
    reason="DRIVE_TEST_FILE_ID env var not set",
)


@pytest.fixture(scope="module")
def service():
    from xlsx_drive_mcp.auth import load_credentials
    from xlsx_drive_mcp.drive import get_drive_service
    creds, _ = load_credentials(None)
    return get_drive_service(creds)


def test_download_upload_roundtrip(service):
    """Download, write a value, upload, re-download, verify the value persisted."""
    from xlsx_drive_mcp.drive import download_xlsx, upload_xlsx

    content, rev_id = download_xlsx(service, DRIVE_TEST_FILE_ID)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    original_value = ws["Z99"].value
    ws["Z99"] = "__integration_test_marker__"
    buf = io.BytesIO()
    wb.save(buf)
    upload_xlsx(service, DRIVE_TEST_FILE_ID, buf.getvalue(), rev_id)

    # Re-download and verify
    content2, _ = download_xlsx(service, DRIVE_TEST_FILE_ID)
    wb2 = openpyxl.load_workbook(io.BytesIO(content2), data_only=True)
    assert wb2.active["Z99"].value == "__integration_test_marker__"

    # Restore original value
    content3, rev_id3 = download_xlsx(service, DRIVE_TEST_FILE_ID)
    wb3 = openpyxl.load_workbook(io.BytesIO(content3))
    wb3.active["Z99"] = original_value
    buf3 = io.BytesIO()
    wb3.save(buf3)
    upload_xlsx(service, DRIVE_TEST_FILE_ID, buf3.getvalue(), rev_id3)


def test_conflict_detection(service):
    """Simulate a concurrent edit: second upload must raise ConflictError."""
    from xlsx_drive_mcp.drive import download_xlsx, upload_xlsx, ConflictError

    content, rev_id = download_xlsx(service, DRIVE_TEST_FILE_ID)

    # First upload succeeds and bumps the revision
    wb = openpyxl.load_workbook(io.BytesIO(content))
    wb.active["Z98"] = "first_write"
    buf = io.BytesIO()
    wb.save(buf)
    upload_xlsx(service, DRIVE_TEST_FILE_ID, buf.getvalue(), rev_id)

    # Second upload with the old rev_id must fail
    wb2 = openpyxl.load_workbook(io.BytesIO(content))
    wb2.active["Z98"] = "second_write"
    buf2 = io.BytesIO()
    wb2.save(buf2)
    with pytest.raises(ConflictError):
        upload_xlsx(service, DRIVE_TEST_FILE_ID, buf2.getvalue(), rev_id)
