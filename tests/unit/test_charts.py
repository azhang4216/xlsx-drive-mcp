import openpyxl
import pytest
from xlsx_drive_mcp.tools.charts import _build_chart, _place_chart, CHART_TYPES


@pytest.fixture
def data_wb():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Month"
    ws["B1"] = "Sales"
    ws["C1"] = "Costs"
    for i, (month, sales, costs) in enumerate([
        ("Jan", 100, 80), ("Feb", 120, 90), ("Mar", 110, 85)
    ], start=2):
        ws[f"A{i}"] = month
        ws[f"B{i}"] = sales
        ws[f"C{i}"] = costs
    return wb


def test_build_bar_col_chart(data_wb):
    chart = _build_chart(data_wb["Sheet1"], "bar", "col", "B2:C4", "A2:A4", ["Sales", "Costs"], "Revenue")
    assert chart is not None
    # openpyxl wraps the title string in a Title object; verify the text is stored
    assert chart.title is not None
    assert chart.title.tx.rich.p[0].r[0].t == "Revenue"


def test_build_line_chart(data_wb):
    chart = _build_chart(data_wb["Sheet1"], "line", "line", "B2:B4", None, None, None)
    assert chart is not None


def test_build_pie_chart(data_wb):
    chart = _build_chart(data_wb["Sheet1"], "pie", "pie", "B2:B4", "A2:A4", None, "Pie")
    assert chart is not None


def test_build_scatter_chart(data_wb):
    chart = _build_chart(data_wb["Sheet1"], "scatter", "marker", "B2:C4", None, None, None)
    assert chart is not None


def test_invalid_chart_type_raises(data_wb):
    with pytest.raises(ValueError, match="chart_type"):
        _build_chart(data_wb["Sheet1"], "funnel", "col", "B2:C4", None, None, None)


def test_place_chart(data_wb):
    chart = _build_chart(data_wb["Sheet1"], "bar", "col", "B2:C4", "A2:A4", None, None)
    _place_chart(data_wb["Sheet1"], chart, "E2")
    assert len(data_wb["Sheet1"]._charts) == 1
