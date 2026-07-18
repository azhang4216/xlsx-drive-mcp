import io
import openpyxl
import pytest
from xlsx_drive_mcp.tools.data import (
    _read_range,
    _write_range,
    _append_rows,
    _clear_range,
    _find_last_data_row,
)


@pytest.fixture
def wb():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["B1"] = "world"
    ws["A2"] = 42
    ws["B2"] = 3.14
    return wb


def test_read_range_basic(wb):
    result = _read_range(wb, "Sheet1", "A1:B2")
    assert result["values"] == [["hello", "world"], [42, 3.14]]
    assert result["cell_range"] == "A1:B2"


def test_read_range_single_cell(wb):
    result = _read_range(wb, "Sheet1", "A1")
    assert result["values"] == [["hello"]]


def test_read_range_unknown_sheet_raises(wb):
    with pytest.raises(KeyError):
        _read_range(wb, "NoSheet", "A1")


def test_write_range_basic(wb):
    _write_range(wb, "Sheet1", "A1", [["new_a", "new_b"], [99, 100]])
    assert wb["Sheet1"]["A1"].value == "new_a"
    assert wb["Sheet1"]["B2"].value == 100


def test_write_range_formula(wb):
    _write_range(wb, "Sheet1", "C1", [["=SUM(A2:B2)"]])
    assert wb["Sheet1"]["C1"].value == "=SUM(A2:B2)"


def test_write_range_does_not_clear_extra_cells(wb):
    """Cells beyond the values array are untouched."""
    _write_range(wb, "Sheet1", "A1", [["only_a1"]])
    assert wb["Sheet1"]["B1"].value == "world"  # unchanged


def test_find_last_data_row_basic(wb):
    assert _find_last_data_row(wb["Sheet1"]) == 2


def test_find_last_data_row_after_clear(wb):
    wb["Sheet1"]["A2"].value = None
    wb["Sheet1"]["B2"].value = None
    assert _find_last_data_row(wb["Sheet1"]) == 1


def test_find_last_data_row_empty():
    wb2 = openpyxl.Workbook()
    assert _find_last_data_row(wb2.active) == 0


def test_append_rows(wb):
    _append_rows(wb, "Sheet1", [["new1", "new2"]])
    assert wb["Sheet1"]["A3"].value == "new1"
    assert wb["Sheet1"]["B3"].value == "new2"


def test_clear_range_preserves_formatting(wb):
    from openpyxl.styles import Font
    wb["Sheet1"]["A1"].font = Font(bold=True)
    _clear_range(wb, "Sheet1", "A1:B1")
    assert wb["Sheet1"]["A1"].value is None
    assert wb["Sheet1"]["A1"].font.bold is True


def test_clear_range_over_merged_raises():
    wb2 = openpyxl.Workbook()
    ws = wb2.active
    ws.title = "Sheet1"
    ws.merge_cells("A1:C1")
    with pytest.raises(ValueError, match="merged"):
        _clear_range(wb2, "Sheet1", "A1:C3")
