import io
import openpyxl
import pytest
from unittest.mock import MagicMock, patch
from xlsx_drive_mcp.tools.files import (
    _build_list_query,
    _get_xlsx_info_from_wb,
    _scan_sheet_dimensions,
)


def test_build_list_query_no_filters():
    q = _build_list_query(None, None)
    assert "spreadsheetml" in q
    assert "name contains" not in q


def test_build_list_query_with_name():
    q = _build_list_query(None, "budget")
    assert "name contains 'budget'" in q


def test_build_list_query_escapes_single_quote():
    q = _build_list_query(None, "Q1'report")
    assert "Q1\\'report" in q


def test_build_list_query_with_folder():
    q = _build_list_query("folder123", None)
    assert "'folder123' in parents" in q


def test_get_xlsx_info_from_wb():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "x"
    ws["C3"] = "y"
    result = _get_xlsx_info_from_wb(wb)
    assert result["sheets"][0]["name"] == "Data"
    assert result["sheets"][0]["rows"] == 3
    assert result["sheets"][0]["cols"] == 3


def test_scan_sheet_dimensions_empty():
    wb = openpyxl.Workbook()
    rows, cols = _scan_sheet_dimensions(wb.active)
    assert rows == 0
    assert cols == 0


def test_scan_sheet_dimensions_after_clear():
    """max_row is unreliable after clear; scan must find actual data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A5"] = "data"
    ws["A5"].value = None  # simulate clear
    rows, cols = _scan_sheet_dimensions(ws)
    assert rows == 0
    assert cols == 0
