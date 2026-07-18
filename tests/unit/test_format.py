import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill, Border, Side
from xlsx_drive_mcp.tools.format import (
    _read_cell_format,
    _apply_format_to_range,
    _parse_color_input,
    _parse_border_input,
    _expand_columns,
    _expand_rows,
)


@pytest.fixture
def wb():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"].font = Font(bold=True, color="FFFF0000", size=14)
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="FF00FF00")
    thin = Side(style="thin")
    ws["C1"].border = Border(top=thin, right=thin)
    return wb


def test_read_cell_format_bold(wb):
    result = _read_cell_format(wb["Sheet1"]["A1"], wb["Sheet1"])
    assert result["bold"] is True
    assert result["font_color"] == "FFFF0000"
    assert result["font_size"] == 14
    assert result["merged"] is False


def test_read_cell_format_bg_color(wb):
    result = _read_cell_format(wb["Sheet1"]["B1"], wb["Sheet1"])
    assert result["bg_color"] == "FF00FF00"


def test_read_cell_format_border(wb):
    result = _read_cell_format(wb["Sheet1"]["C1"], wb["Sheet1"])
    assert result["border"]["top"] == "thin"
    assert result["border"]["right"] == "thin"
    assert result["border"]["left"] is None


def test_parse_color_six_digit():
    assert _parse_color_input("FF0000") == "FFFF0000"


def test_parse_color_eight_digit():
    assert _parse_color_input("FFFF0000") == "FFFF0000"


def test_parse_color_invalid():
    with pytest.raises(ValueError):
        _parse_color_input("ZZZZZZ")


def test_parse_border_string_shorthand():
    result = _parse_border_input("thin")
    assert result == {"top": "thin", "bottom": "thin", "left": "thin", "right": "thin"}


def test_parse_border_dict():
    result = _parse_border_input({"top": "medium", "bottom": "none"})
    assert result["top"] == "medium"
    assert result["bottom"] == "none"
    assert result.get("left") is None


def test_expand_columns_letter():
    assert _expand_columns("A") == ["A"]


def test_expand_columns_list():
    assert _expand_columns(["A", "C"]) == ["A", "C"]


def test_expand_columns_range_string():
    assert _expand_columns("A:C") == ["A", "B", "C"]


def test_expand_rows_single():
    assert _expand_rows(3) == [3]


def test_expand_rows_list():
    assert _expand_rows([1, 3, 5]) == [1, 3, 5]


def test_expand_rows_range_string():
    assert _expand_rows("2:4") == [2, 3, 4]


def test_apply_format_bold(wb):
    _apply_format_to_range(wb["Sheet1"], "A2:B2", bold=True)
    assert wb["Sheet1"]["A2"].font.bold is True
    assert wb["Sheet1"]["B2"].font.bold is True
