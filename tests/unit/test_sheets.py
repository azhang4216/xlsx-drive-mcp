import io
import openpyxl
import pytest
from xlsx_drive_mcp.tools.sheets import (
    _list_sheets,
    _create_sheet,
    _delete_sheet,
    _rename_sheet,
    _copy_sheet,
)


@pytest.fixture
def multi_wb():
    wb = openpyxl.Workbook()
    wb.active.title = "Alpha"
    wb.create_sheet("Beta")
    return wb


def test_list_sheets(multi_wb):
    result = _list_sheets(multi_wb)
    assert result == [{"name": "Alpha", "index": 0}, {"name": "Beta", "index": 1}]


def test_create_sheet_at_end(multi_wb):
    result = _create_sheet(multi_wb, "Gamma", None)
    assert result["name"] == "Gamma"
    assert result["index"] == 2
    assert "Gamma" in multi_wb.sheetnames


def test_create_sheet_at_position(multi_wb):
    _create_sheet(multi_wb, "First", 0)
    assert multi_wb.sheetnames[0] == "First"


def test_delete_sheet(multi_wb):
    _delete_sheet(multi_wb, "Beta")
    assert "Beta" not in multi_wb.sheetnames


def test_delete_last_sheet_raises(multi_wb):
    _delete_sheet(multi_wb, "Beta")
    with pytest.raises(ValueError, match="only sheet"):
        _delete_sheet(multi_wb, "Alpha")


def test_rename_sheet(multi_wb):
    _rename_sheet(multi_wb, "Beta", "Delta")
    assert "Delta" in multi_wb.sheetnames
    assert "Beta" not in multi_wb.sheetnames


def test_copy_sheet(multi_wb):
    result = _copy_sheet(multi_wb, "Alpha", "AlphaCopy")
    assert result["name"] == "AlphaCopy"
    assert "AlphaCopy" in multi_wb.sheetnames


def test_copy_sheet_nonexistent_raises(multi_wb):
    with pytest.raises(KeyError):
        _copy_sheet(multi_wb, "NoSuchSheet", "Copy")
